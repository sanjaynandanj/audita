"""XLSX export of a recon report — the artifact a CFO forwards to their CA."""

from __future__ import annotations

import io
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Font

from .builder import Report

_HEADERS = [
    "Exception ID", "Bucket", "ITC Amount (₹)", "Reason",
    "Supplier GSTIN", "Supplier Name", "Invoice No (Books)", "Invoice No (2B)",
    "Invoice Date", "Taxable (Books)", "Taxable (2B)", "Tax (Books)", "Tax (2B)",
    "Verified", "Verified By", "CA Sign-off", "Source Ref",
]


def _row(item) -> list:
    books, g2b = item.books or {}, item.gstr2b or {}
    return [
        item.exception_id,
        item.bucket,
        float(Decimal(item.itc_amount)),
        item.reason,
        books.get("gstin") or g2b.get("gstin", ""),
        books.get("supplier_name") or g2b.get("supplier_name", ""),
        books.get("invoice_no", ""),
        g2b.get("invoice_no", ""),
        books.get("invoice_date") or g2b.get("invoice_date", ""),
        books.get("taxable_value", ""),
        g2b.get("taxable_value", ""),
        books.get("total_tax", ""),
        g2b.get("total_tax", ""),
        "YES" if item.verified else "PENDING",
        item.verified_by,
        item.ca_signoff,
        books.get("source_ref") or g2b.get("source_ref", ""),
    ]


def export_xlsx(report: Report) -> bytes:
    wb = Workbook()
    bold = Font(bold=True)

    summary = wb.active
    summary.title = "Summary"
    rows = [
        ("Audita — GST ITC Reconciliation Report", ""),
        ("Client", report.client_name),
        ("Generated", report.created_at),
        ("Period", report.period_note),
        ("", ""),
        ("ITC AT RISK (verified)", f"₹{report.verified_at_risk}"),
        ("ITC at risk (pending verification)", f"₹{report.pending_at_risk}"),
        ("Missed ITC (in 2B, not booked)", f"₹{report.missed_itc_total}"),
        ("Unresolved (needs review)", f"₹{report.unresolved_total}"),
        ("Matched invoices", str(report.matched_count)),
        ("Matched tax total", f"₹{report.matched_tax_total}"),
        ("", ""),
        ("Headline counts only human-verified exceptions.", ""),
        ("Unresolved items (credit/debit notes, amendments, RCM, ISD,", ""),
        ("ambiguous matches) are listed separately and never in the headline.", ""),
    ]
    for r in rows:
        summary.append(r)
    summary["A1"].font = bold
    summary["A6"].font = bold
    summary.column_dimensions["A"].width = 45
    summary.column_dimensions["B"].width = 30

    for title, items in (
        ("ITC At Risk", report.exceptions),
        ("Missed ITC", report.missed_itc),
        ("Unresolved", report.unresolved),
    ):
        ws = wb.create_sheet(title)
        ws.append(_HEADERS)
        for cell in ws[1]:
            cell.font = bold
        for item in items:
            ws.append(_row(item))
        ws.column_dimensions["D"].width = 40

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
