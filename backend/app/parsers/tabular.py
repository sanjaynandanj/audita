"""Shared CSV/XLSX row parser with flexible column mapping.

Used for the purchase register (Tally exports vary wildly) and the
simplified GSTR-2B tabular fallback.
"""

from __future__ import annotations

import csv
from decimal import Decimal, InvalidOperation
from pathlib import Path

from ..engine.models import InvoiceRecord, Source

# canonical field -> accepted header aliases (lowercased, stripped)
COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "gstin": ("gstin", "gstin of supplier", "supplier gstin", "gstin/uin", "gstin no",
              "party gstin", "ctin"),
    "invoice_no": ("invoice no", "invoice number", "invoice_no", "inv no", "bill no", "voucher no",
                   "document number", "inum", "supplier invoice no"),
    "invoice_date": ("invoice date", "date", "inv date", "bill date", "voucher date",
                     "document date", "dt"),
    "supplier_name": ("supplier name", "supplier", "party name", "party", "name of supplier",
                      "trade name", "trdnm", "particulars"),
    "taxable_value": ("taxable value", "taxable amount", "taxable", "txval", "assessable value", "base amount"),
    "igst": ("igst", "igst amount", "integrated tax", "iamt"),
    "cgst": ("cgst", "cgst amount", "central tax", "camt"),
    "sgst": ("sgst", "sgst amount", "state tax", "state/ut tax", "samt"),
    "cess": ("cess", "cess amount", "csamt"),
    "doc_type": ("doc type", "document type", "type", "note type"),
    "reverse_charge": ("reverse charge", "rcm", "rev", "reverse charge applicable"),
}


def _dec(value) -> Decimal:
    if value is None:
        return Decimal("0")
    text = str(value).strip().replace(",", "").replace("₹", "")
    if not text or text in ("-", "--"):
        return Decimal("0")
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return Decimal("0")


def _build_column_map(headers: list[str]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    normalized = [str(h or "").strip().lower().replace(".", "").replace("_", " ") for h in headers]
    for field, aliases in COLUMN_ALIASES.items():
        for idx, header in enumerate(normalized):
            if header in aliases and field not in mapping:
                mapping[field] = idx
                break
    return mapping


def _rows_from_csv(path: Path) -> list[list]:
    with open(path, newline="", encoding="utf-8-sig") as fh:
        return [row for row in csv.reader(fh)]


def _rows_from_xlsx(path: Path) -> list[list]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = [[cell for cell in row] for row in ws.iter_rows(values_only=True)]
    wb.close()
    return rows


def parse_tabular(path: str | Path, source: Source) -> list[InvoiceRecord]:
    path = Path(path)
    suffix = path.suffix.lower()
    if suffix == ".csv":
        rows = _rows_from_csv(path)
    elif suffix in (".xlsx", ".xls"):
        rows = _rows_from_xlsx(path)
    else:
        raise ValueError(f"Unsupported tabular file type: {suffix}")

    if not rows:
        return []

    # Find the header row: first row where we can map both gstin and invoice_no
    header_idx = None
    colmap: dict[str, int] = {}
    for idx, row in enumerate(rows[:10]):
        candidate = _build_column_map([str(c) if c is not None else "" for c in row])
        if "gstin" in candidate and "invoice_no" in candidate:
            header_idx, colmap = idx, candidate
            break
    if header_idx is None:
        raise ValueError(
            f"Could not locate a header row with GSTIN and Invoice No columns in {path.name}. "
            f"Accepted GSTIN headers: {', '.join(COLUMN_ALIASES['gstin'])}"
        )

    def cell(row: list, field: str) -> str:
        idx = colmap.get(field)
        if idx is None or idx >= len(row) or row[idx] is None:
            return ""
        return str(row[idx]).strip()

    records: list[InvoiceRecord] = []
    for rownum, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
        gstin = cell(row, "gstin")
        invoice_no = cell(row, "invoice_no")
        if not gstin and not invoice_no:
            continue  # blank/total row
        doc_type_raw = cell(row, "doc_type").upper()
        doc_type = "CDN" if doc_type_raw in ("C", "D", "CDN", "CREDIT NOTE", "DEBIT NOTE", "CR", "DR") else "INV"
        rcm_raw = cell(row, "reverse_charge").upper()
        records.append(
            InvoiceRecord(
                source=source,
                gstin=gstin,
                invoice_no=invoice_no,
                invoice_date=cell(row, "invoice_date"),
                supplier_name=cell(row, "supplier_name"),
                taxable_value=_dec(cell(row, "taxable_value")),
                igst=_dec(cell(row, "igst")),
                cgst=_dec(cell(row, "cgst")),
                sgst=_dec(cell(row, "sgst")),
                cess=_dec(cell(row, "cess")),
                doc_type=doc_type,
                reverse_charge=rcm_raw in ("Y", "YES", "TRUE", "1"),
                source_ref=f"{path.name}:row{rownum}",
            )
        )
    return records
